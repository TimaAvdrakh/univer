from rest_framework import generics, views
from rest_framework import status
from rest_framework.response import Response
from . import serializers
from organizations import models as org_models
from common.serializers import AcadPeriodSerializer
from common import models as common_models
from portal_users.serializers import EducationProgramSerializer, EducationProgramGroupSerializer, \
    StudyPlanSerializer, ProfileShortSerializer
from portal_users.models import Profile
from organizations.models import StudentDisciplineInfo
from portal.curr_settings import student_discipline_info_status, student_discipline_status, STUDENT_STATUSES
from django.db.models import Q, Count, Sum
from common.paginators import CustomPagination, AdvisorBidPagination
from . import permissions as adv_permission
from rest_framework.permissions import IsAuthenticated
from . import models
from django.db.models import Q, Count, Max
from openpyxl import Workbook, load_workbook
from datetime import datetime
from django.shortcuts import HttpResponse
from uuid import uuid4
from portal_users.utils import get_current_study_year
from openpyxl.styles import Border, Side, Font, Alignment
from rest_framework.views import APIView
from django.db import connection
from portal.curr_settings import current_site
from cron_app.models import ExcelTask
from django.core.cache import cache
from django.db.models import Value, F, Subquery
from django.db.models.functions import Concat
import json
from django.db.models import Q


class StudyPlansListView(generics.ListAPIView):
    """
    Получение учебных планов,
    study_year(!), study_form, faculty, cathedra, edu_prog_group, edu_prog, course, group, status
    """
    queryset = org_models.StudyPlan.objects.filter(is_active=True).order_by('student__last_name')
    serializer_class = serializers.StudyPlanSerializer

    # pagination_class = CustomPagination

    def get_queryset(self):
        study_year = self.request.query_params.get('study_year')  # Дисциплина студента
        study_form = self.request.query_params.get('study_form')
        faculty = self.request.query_params.get('faculty')
        cathedra = self.request.query_params.get('cathedra')
        edu_prog_group = self.request.query_params.get('edu_prog_group')
        edu_prog = self.request.query_params.get('edu_prog')
        course = self.request.query_params.get('course')  # Дисциплина студента
        group = self.request.query_params.get('group')

        status_id = self.request.query_params.get('status')  # В Дисциплине студента
        # reg_period = self.request.query_params.get('reg_period')  # Дисциплина студента

        queryset = self.queryset.filter(advisor=self.request.user.profile)

        if status_id:
            status_obj = org_models.StudentDisciplineStatus.objects.get(number=status_id)

            study_plan_pks_from_sd = org_models.StudentDiscipline.objects.filter(
                study_year_id=study_year,
                status=status_obj,
            ).values('study_plan')

            queryset = queryset.filter(pk__in=study_plan_pks_from_sd)

        if study_form:
            queryset = queryset.filter(study_form_id=study_form)
        if faculty:
            queryset = queryset.filter(faculty_id=faculty)
        if cathedra:
            queryset = queryset.filter(cathedra_id=cathedra)
        if edu_prog:
            queryset = queryset.filter(education_program_id=edu_prog)
        if edu_prog_group:
            queryset = queryset.filter(education_program__group_id=edu_prog_group)
        if group:
            queryset = queryset.filter(group_id=group)
        if study_year:
            study_year_obj = org_models.StudyPeriod.objects.get(pk=study_year)
            queryset = queryset.filter(study_period__end__gt=study_year_obj.start)
        if course and study_year:
            study_plan_pks = org_models.StudyYearCourse.objects.filter(
                study_year_id=study_year,
                course=course
            ).values('study_plan')
            queryset = queryset.filter(pk__in=study_plan_pks)

        return queryset


class StudentDisciplineListView(generics.ListAPIView):
    """
    Получение дисциплин студента, query_params:
    study_plan(!), study_year(!), acad_period(!), status, short(!) (если значение 1, вернет только первые три записи)
    """
    queryset = org_models.StudentDiscipline.objects.filter(is_active=True)
    serializer_class = serializers.StudentDisciplineShortSerializer

    def list(self, request, *args, **kwargs):
        short = self.request.query_params.get('short')

        study_plan = self.request.query_params.get('study_plan')
        study_year = self.request.query_params.get('study_year')
        acad_period = self.request.query_params.get('acad_period')
        status_id = self.request.query_params.get('status')
        # reg_period = self.request.query_params.get('reg_period')

        checks = models.AdvisorCheck.objects.filter(study_plan_id=study_plan,
                                                    acad_period_id=acad_period)

        if checks.exists():
            old_status = checks.latest('id').status
        else:
            old_status = 0

        queryset = self.queryset.all()
        if study_plan:
            queryset = queryset.filter(study_plan_id=study_plan)
        if status_id:
            status_obj = org_models.StudentDisciplineStatus.objects.get(number=status_id)
            queryset = queryset.filter(status=status_obj)
        if study_year:
            queryset = queryset.filter(study_year_id=study_year)
        if acad_period:
            queryset = queryset.filter(acad_period_id=acad_period)

        queryset = queryset.distinct('discipline').order_by('discipline')  # TODO TEST
        total_credit = sum(i.credit for i in queryset)

        is_more = len(queryset) > 3

        if int(short) == 1:
            queryset = queryset[:3]
        else:
            is_more = False

        serializer = self.serializer_class(instance=queryset,
                                           many=True)
        try:
            acad_period_obj = org_models.AcadPeriod.objects.get(pk=acad_period,
                                                            is_active=True).repr_name
        except:
            return Response({'acad_period_obj': None}, status=status.HTTP_404_NOT_FOUND)

        resp = {
            'total_credit': total_credit,
            'disciplines': serializer.data,
            'is_more': is_more,
            'old_status': old_status,
            'acad_period': acad_period_obj,
            'uid': acad_period,
        }

        return Response(
            resp,
            status=status.HTTP_200_OK
        )


class StudentDisciplineGroupListView(views.APIView):  # TODO CHECK
    """
    Получение дисциплин студента группой, query_params:
    study_plan(!), study_year(!), acad_period(!), status
    """
    queryset = org_models.StudentDiscipline.objects.filter(is_active=True)
    serializer_class = serializers.StudentDisciplineShortSerializer

    def post(self, request, *args, **kwargs):
        study_plans = self.request.data.get('study_plans')
        study_year = self.request.data.get('study_year')
        acad_period = self.request.data.get('acad_period')
        status_id = self.request.data.get('status')
        # reg_period = self.request.query_params.get('reg_period')

        study_plan_list = study_plans.split(',')

        disciplines = []

        for sp in study_plan_list:

            checks = models.AdvisorCheck.objects.filter(study_plan_id=sp,
                                                        acad_period_id=acad_period)

            if checks.exists():
                old_status = checks.latest('id').status
            else:
                old_status = 0

            queryset = self.queryset.all()
            if sp:
                queryset = queryset.filter(study_plan_id=sp)
            if status_id:
                status_obj = org_models.StudentDisciplineStatus.objects.get(number=status_id)
                queryset = queryset.filter(status=status_obj)
            if study_year:
                queryset = queryset.filter(study_year_id=study_year)
            if acad_period:
                queryset = queryset.filter(acad_period_id=acad_period)

            queryset = queryset.distinct('discipline').order_by('discipline')  # TODO TEST
            total_credit = sum(i.credit for i in queryset)

            is_more = len(queryset) > 3
            queryset = queryset[:3]
            serializer = self.serializer_class(instance=queryset,
                                               many=True)

            acad_period_obj = org_models.AcadPeriod.objects.get(pk=acad_period,
                                                                is_active=True)

            resp = {
                'study_plan': sp,
                'total_credit': total_credit,
                'disciplines': serializer.data,
                'is_more': is_more,
                'old_status': old_status,
                'acad_period': acad_period_obj.repr_name,
                'uid': acad_period,
            }

            disciplines.append(resp)

        return Response(
            disciplines,
            status=status.HTTP_200_OK
        )


class AcadPeriodListView(generics.ListAPIView):
    """Получить список акад периодов по курсу и периоду регистрации,
    study_year(!) reg_period(!), study_form, faculty, cathedra, edu_prog_group, edu_prog, course, group, status
    """

    queryset = org_models.AcadPeriod.objects.filter(is_active=True).order_by('number')
    serializer_class = AcadPeriodSerializer

    def get_queryset(self):
        reg_period = self.request.query_params.get('reg_period')
        study_year = self.request.query_params.get('study_year')

        course = self.request.query_params.get('course')
        study_form = self.request.query_params.get('study_form')
        faculty = self.request.query_params.get('faculty')
        cathedra = self.request.query_params.get('cathedra')
        edu_prog_group = self.request.query_params.get('edu_prog_group')
        edu_prog = self.request.query_params.get('edu_prog')
        group = self.request.query_params.get('group')
        status_id = self.request.query_params.get('status')

        # profile = self.request.user.profile

        if course:
            acad_period_pks = common_models.CourseAcadPeriodPermission.objects.filter(
                registration_period_id=reg_period,
                course=course
            ).values('acad_period')
        else:
            acad_period_pks = common_models.CourseAcadPeriodPermission.objects.filter(
                registration_period_id=reg_period,
            ).values('acad_period')

        acad_periods = self.queryset.filter(pk__in=acad_period_pks)

        sd = org_models.StudentDiscipline.objects.filter(
            # study_plan__advisor=profile,
            study_year_id=study_year,
        )

        if study_form:
            sd = sd.filter(study_plan__study_form_id=study_form)
        if faculty:
            sd = sd.filter(study_plan__faculty_id=faculty)
        if cathedra:
            sd = sd.filter(study_plan__cathedra_id=cathedra)
        if edu_prog_group:
            sd = sd.filter(study_plan__education_program__group_id=edu_prog_group)
        if edu_prog:
            sd = sd.filter(study_plan__education_program_id=edu_prog)
        if group:
            sd = sd.filter(study_plan__group_id=group)
        if status_id:
            status_obj = org_models.StudentDisciplineStatus.objects.get(number=status_id)
            sd = sd.filter(status=status_obj)

        acad_period_pks_from_sd = sd.values('acad_period')

        acad_periods = acad_periods.filter(pk__in=acad_period_pks_from_sd)

        return acad_periods


class FacultyListView(generics.ListAPIView):
    """Получить список факультетов доступных для эдвайзеру,
    study_year(!), study_form"""
    queryset = org_models.Faculty.objects.filter(is_active=True).order_by('name')
    serializer_class = serializers.FacultySerializer

    def get_queryset(self):
        profile = self.request.user.profile

        study_year = self.request.query_params.get('study_year')
        study_form = self.request.query_params.get('study_form')
        study_plans = org_models.StudyPlan.objects.filter(advisor=profile)

        if study_form:
            study_plans = study_plans.filter(study_form_id=study_form)
        if study_year:
            study_year_obj = org_models.StudyPeriod.objects.get(pk=study_year)
            study_plans = study_plans.filter(study_period__end__gt=study_year_obj.start)

        faculty_pks = study_plans.values('faculty')

        queryset = self.queryset.filter(pk__in=faculty_pks)
        return queryset


class CathedraListView(generics.ListAPIView):
    """Получить список кафедр доступных для эдвайзеру,
    study_year, study_form, faculty"""

    queryset = org_models.Cathedra.objects.filter(is_active=True).order_by('name')
    serializer_class = serializers.CathedraSerializer

    def get_queryset(self):
        profile = self.request.user.profile

        study_year = self.request.query_params.get('study_year')
        study_form = self.request.query_params.get('study_form')
        faculty = self.request.query_params.get('faculty')

        study_plans = org_models.StudyPlan.objects.filter(advisor=profile)

        if study_form:
            study_plans = study_plans.filter(study_form_id=study_form)
        if study_year:
            study_year_obj = org_models.StudyPeriod.objects.get(pk=study_year)
            study_plans = study_plans.filter(study_period__end__gt=study_year_obj.start)

        cathedra_pks = study_plans.values('cathedra')
        queryset = self.queryset.filter(pk__in=cathedra_pks)

        if faculty:
            queryset = queryset.filter(faculty_id=faculty)

        return queryset


class EducationProgramGroupListView(generics.ListAPIView):
    """Получить список групп образовательных программ.
    study_year, study_form, faculty, cathedra, speciality(для 2 стр (Утвержденные дисциплины))"""

    queryset = org_models.EducationProgramGroup.objects.filter(is_active=True).order_by('name')
    serializer_class = EducationProgramGroupSerializer

    def get_queryset(self):
        profile = self.request.user.profile

        study_year = self.request.query_params.get('study_year')
        study_form = self.request.query_params.get('study_form')
        faculty = self.request.query_params.get('faculty')
        cathedra = self.request.query_params.get('cathedra')
        speciality = self.request.query_params.get('speciality')  # Для утвержденных дисциплин

        study_plans = org_models.StudyPlan.objects.filter(advisor=profile,
                                                          is_active=True)

        if study_form:
            study_plans = study_plans.filter(study_form_id=study_form)
        if study_year:
            study_year_obj = org_models.StudyPeriod.objects.get(pk=study_year)
            study_plans = study_plans.filter(study_period__end__gt=study_year_obj.start)
        if faculty:
            study_plans = study_plans.filter(faculty_id=faculty)
        if cathedra:
            study_plans = study_plans.filter(cathedra_id=cathedra)
        if speciality:  # Для утвержденных дисциплин
            study_plans = study_plans.filter(speciality_id=speciality)

        education_program_pks = study_plans.values('education_program')
        queryset = self.queryset.filter(educationprogram__in=education_program_pks)

        return queryset


class EducationProgramListView(generics.ListAPIView):
    """Получить список образовательных программ,
    edu_prog_group, study_year, study_form, faculty, cathedra"""

    queryset = org_models.EducationProgram.objects.filter(is_active=True).order_by('name')
    serializer_class = EducationProgramSerializer

    def get_queryset(self):
        profile = self.request.user.profile

        study_year = self.request.query_params.get('study_year')
        study_form = self.request.query_params.get('study_form')
        edu_prog_group = self.request.query_params.get('edu_prog_group')
        cathedra = self.request.query_params.get('cathedra')
        faculty = self.request.query_params.get('faculty')
        speciality = self.request.query_params.get('speciality')

        study_plans = org_models.StudyPlan.objects.filter(advisor=profile,
                                                          is_active=True)

        if study_form:
            study_plans = study_plans.filter(study_form_id=study_form)
        if study_year:
            study_year_obj = org_models.StudyPeriod.objects.get(pk=study_year)
            study_plans = study_plans.filter(study_period__end__gt=study_year_obj.start)
        if faculty:
            study_plans = study_plans.filter(faculty_id=faculty)
        if cathedra:
            study_plans = study_plans.filter(cathedra_id=cathedra)
        if speciality:
            study_plans = study_plans.filter(speciality_id=speciality)

        education_program_pks = study_plans.values('education_program')
        queryset = self.queryset.filter(pk__in=education_program_pks)

        if edu_prog_group:
            queryset = queryset.filter(group_id=edu_prog_group)
        return queryset


class GroupListView(generics.ListAPIView):
    """Получить список групп,
    study_year, study_form, faculty, cathedra, edu_prog_group, edu_prog, course"""

    queryset = org_models.Group.objects.filter(is_active=True).order_by('name')
    serializer_class = serializers.GroupShortSerializer

    def get_queryset(self):
        profile = self.request.user.profile

        study_year = self.request.query_params.get('study_year')
        study_form = self.request.query_params.get('study_form')
        cathedra = self.request.query_params.get('cathedra')
        faculty = self.request.query_params.get('faculty')

        edu_prog = self.request.query_params.get('edu_prog')
        edu_prog_group = self.request.query_params.get('edu_prog_group')
        course = self.request.query_params.get('course')
        speciality = self.request.query_params.get('speciality')  # Для утвержденных дисциплин

        study_plans = org_models.StudyPlan.objects.filter(advisor=profile)

        if study_form:
            study_plans = study_plans.filter(study_form_id=study_form)

        if edu_prog:
            study_plans = study_plans.filter(education_program_id=edu_prog)
        elif edu_prog_group:
            study_plans = study_plans.filter(education_program__group_id=edu_prog_group)

        if course and study_year:
            study_plan_pks = org_models.StudyYearCourse.objects.filter(
                study_year_id=study_year,
                course=course
            ).values('study_plan')

            study_plans = study_plans.filter(pk__in=study_plan_pks)

        if study_year:
            study_year_obj = org_models.StudyPeriod.objects.get(pk=study_year)
            study_plans = study_plans.filter(study_period__end__gt=study_year_obj.start)
        if faculty:
            study_plans = study_plans.filter(faculty_id=faculty)
        if cathedra:
            study_plans = study_plans.filter(cathedra_id=cathedra)

        if speciality:  # Для утвержденных дисциплин
            study_plans = study_plans.filter(speciality_id=speciality)

        group_pks = study_plans.values('group')
        queryset = self.queryset.filter(pk__in=group_pks)

        return queryset


class CheckStudentChoices(generics.CreateAPIView):
    """Утвердить или отклонить заявки студента. Статус: 4 - утврежден, 3 - отклонен"""
    serializer_class = serializers.CheckStudentBidsSerializer
    permission_classes = (
        IsAuthenticated,
        adv_permission.StudyPlanAdvisorPermission
    )

    def create(self, request, *args, **kwargs):
        try:
            sp = org_models.StudyPlan.objects.get(pk=request.data['study_plan'])
        except org_models.StudyPlan.DoesNotExist:
            return Response(
                {
                    "message": "not_found",
                },
                status=status.HTTP_404_NOT_FOUND
            )
        self.check_object_permissions(request, sp)

        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(
            {
                'message': 'ok'
            },
            status=status.HTTP_200_OK
        )


class FilteredStudentsListView(generics.ListAPIView):
    """Фильтровать студентов для селекта (Эдвайзер)
        study_year(!), edu_prog(!), edu_prog_group, faculty, speciality, group"""
    serializer_class = ProfileShortSerializer

    def get_queryset(self):
        request = self.request
        profile = request.user.profile

        study_year = request.query_params.get('study_year')
        faculty = request.query_params.get('faculty')
        speciality = request.query_params.get('speciality')
        edu_prog = request.query_params.get('edu_prog')
        edu_prog_group = self.request.query_params.get('edu_prog_group')
        group = request.query_params.get('group')

        study_year_obj = org_models.StudyPeriod.objects.get(pk=study_year)

        study_plans = org_models.StudyPlan.objects.filter(
            study_period__end__gt=study_year_obj.start,
            advisor=profile,
            is_active=True,
        )

        if edu_prog:
            study_plans = study_plans.filter(education_program_id=edu_prog)
        elif edu_prog_group:
            study_plans = study_plans.filter(education_program__group_id=edu_prog_group)

        if faculty:
            study_plans = study_plans.filter(faculty_id=faculty)

        if speciality:
            study_plans = study_plans.filter(speciality_id=speciality)

        if group:
            study_plans = study_plans.filter(group_id=group)

        student_pks = study_plans.values('student')
        students = Profile.objects.filter(pk__in=student_pks).\
            exclude(status_id=STUDENT_STATUSES['expelled']).order_by('last_name')

        return students


class SpecialityListView(generics.ListAPIView):
    """Получить список специальностей доступных эдвайзеру,
    study_year(!), faculty"""

    queryset = org_models.Speciality.objects.filter(is_active=True).order_by('name')
    serializer_class = serializers.SpecialitySerializer

    def get_queryset(self):
        profile = self.request.user.profile

        study_year = self.request.query_params.get('study_year')
        faculty = self.request.query_params.get('faculty')
        cathedra = self.request.query_params.get('cathedra')

        study_plans = org_models.StudyPlan.objects.filter(advisor=profile,
                                                          is_active=True)

        if faculty:
            study_plans = study_plans.filter(faculty_id=faculty)

        if cathedra:
            study_plans = study_plans.filter(cathedra_id=cathedra)

        if study_year:
            study_year_obj = org_models.StudyPeriod.objects.get(pk=study_year)
            study_plans = study_plans.filter(study_period__end__gt=study_year_obj.start)

        speciality_pks = study_plans.values('speciality')
        queryset = self.queryset.filter(pk__in=speciality_pks)

        return queryset


class GetStudyPlanView(generics.RetrieveAPIView):
    """Получить учебный план студента для ИУПы обучающихся (2 стр, Утвержденные)
        study_year(!), edu_prog(!), student(!), faculty, speciality, group
    """
    serializer_class = serializers.StudyPlanDetailSerializer

    def get(self, request, *args, **kwargs):
        profile = request.user.profile

        study_year = request.query_params.get('study_year')
        faculty = request.query_params.get('faculty')
        speciality = request.query_params.get('speciality')
        edu_prog = request.query_params.get('edu_prog')
        group = request.query_params.get('group')
        student = request.query_params.get('student')

        filters = {}

        if faculty:
            filters['faculty_id'] = faculty

        if speciality:
            filters['speciality_id'] = speciality

        if group:
            filters['group_id'] = group

        try:
            study_year_obj = org_models.StudyPeriod.objects.get(pk=study_year)
            study_plan = org_models.StudyPlan.objects.get(
                advisor=profile,
                study_period__end__gt=study_year_obj.start,
                education_program_id=edu_prog,
                student_id=student,
                is_active=True,
                **filters,
            )
        except org_models.StudyPlan.DoesNotExist:
            return Response(
                {
                    'message': 0  # Учебный план не найден
                },
                status=status.HTTP_404_NOT_FOUND
            )

        serializer = self.serializer_class(instance=study_plan,
                                           context={'study_year_obj': study_year_obj})
        return Response(
            serializer.data,
            status=status.HTTP_200_OK,
        )


class ConfirmedStudentDisciplineListView(generics.ListAPIView):
    """Получить все утвержденные дисциплины выбранного учебного плана
        study_plan(!), study_year(!), reg_period(!)"""

    serializer_class = serializers.ConfirmedStudentDisciplineShortSerializer

    def list(self, request, *args, **kwargs):
        study_plan_id = request.query_params.get('study_plan')
        study_year = request.query_params.get('study_year')
        reg_period = self.request.query_params.get('reg_period')

        study_plan = org_models.StudyPlan.objects.get(pk=study_plan_id,
                                                      is_active=True)
        study_year_obj = org_models.StudyPeriod.objects.get(pk=study_year)
        course = study_plan.get_course(study_year_obj)

        acad_period_pks_from_sd = org_models.StudentDiscipline.objects.filter(
            study_year_id=study_year,
            study_plan_id=study_plan_id,
            is_active=True,
        ).distinct('acad_period').values('acad_period')

        acad_periods = org_models.AcadPeriod.objects.filter(pk__in=acad_period_pks_from_sd)

        acad_period_pks_from_rule = common_models.CourseAcadPeriodPermission.objects.filter(
            registration_period_id=reg_period,
            course=course
        ).values('acad_period')

        acad_periods = acad_periods.filter(pk__in=acad_period_pks_from_rule)

        resp = []

        for acad_period in acad_periods:
            if StudentDisciplineInfo.objects.filter(
                    study_plan_id=study_plan_id,
                    acad_period=acad_period,
                    status_id=student_discipline_info_status['confirmed'],
            ).exists():
                student_disciplines = org_models.StudentDiscipline.objects.filter(
                    study_year_id=study_year,
                    study_plan_id=study_plan_id,
                    acad_period=acad_period,
                ).distinct('discipline').order_by('discipline')  # TODO TEST

                serializer = self.serializer_class(student_disciplines,
                                                   many=True)
                item = {
                    'acad_period': acad_period.repr_name,
                    'disciplines': serializer.data,
                }
                resp.append(item)

        return Response(
            resp,
            status=status.HTTP_200_OK
        )


class RegisterResultView(generics.ListAPIView):
    """Результат регистрации
        study_year(!), reg_period(!), acad_period, faculty, speciality, edu_prog, course, group
    """
    serializer_class = serializers.StudentDisciplineSerializer
    page_size = 30
    queryset = org_models.StudentDiscipline.objects.exclude(
        student__status_id=STUDENT_STATUSES['expelled'],
        is_active=False)
    pagination_class = CustomPagination

    def list(self, request, *args, **kwargs):
        query = dict()
        queryset = self.queryset
        my = request.query_params.get('my')

        query = dict()

        profile = request.user.profile

        if my == '1':  # Если пользователь является учителем
            queryset = queryset.filter(teacher=profile)
        else:
            query['study_plan__advisor'] = profile

        query['status_id'] = student_discipline_status['confirmed']

        if request.query_params.get('acad_period'):
            query['acad_period_id'] = request.query_params.get('acad_period')

        if request.query_params.get('reg_period'):
            query['acad_period__in'] = common_models.CourseAcadPeriodPermission.objects.filter(
                registration_period_id=request.query_params.get('reg_period'),
            ).values('acad_period')

        if request.query_params.get('faculty'):
            query['study_plan__faculty_id'] = request.query_params.get('faculty')

        if request.query_params.get('speciality'):
            query['study_plan__speciality_id'] = request.query_params.get('speciality')

        if request.query_params.get('edu_prog'):
            query['study_plan__education_program_id'] = request.query_params.get('edu_prog')

        if request.query_params.get('group'):
            query['study_plan__group_id'] = request.query_params.get('group')

        if request.query_params.get('study_year'):
            query['study_year_id'] = request.query_params.get('study_year')

        study_year = request.query_params.get('study_year')
        course = request.query_params.get('course')

        if course and study_year:
            query['study_plan__in'] = org_models.StudyYearCourse.objects.filter(
                study_year_id=study_year,
                course=course
            ).values('study_plan')

        distincted_queryset = queryset.filter(**query).distinct(
            'discipline', 'load_type', 'hours', 'language', 'teacher')

        if request.query_params.getlist('ordering[]'):
            distincted_queryset = queryset.filter(
                uid__in=Subquery(distincted_queryset.values_list('uid', flat=True))
            ).order_by(*request.query_params.getlist('ordering[]'))

        page = self.paginate_queryset(distincted_queryset)
        # print(page, 'page')

        student_discipline_list = []
        for item in page:
            student_count = queryset.filter(**query).distinct('student').count()

            teacher = ''
            if item.teacher:
                teacher = item.teacher.full_name

            d = {
                'discipline_id': item.discipline_id,
                'discipline': item.discipline.name,
                'load_type': item.load_type.name,
                'hours': item.hours,
                'language': item.language.name,
                'teacher': teacher,
                'student_count': student_count,
                'load_type_id': item.load_type_id,
                'language_id': item.language_id,
            }
            student_discipline_list.append(d)

        # page = self.paginate_queryset(student_discipline_list)
        if student_discipline_list is not None:
            # serializer = self.serializer_class(student_discipline_list, many=True)
            return self.get_paginated_response(student_discipline_list)


class RegisterStatisticsView(generics.ListAPIView):
    """Статистика регистрации
    study_year(!), reg_period(!), acad_period, faculty, speciality, edu_prog, course, group
    """
    serializer_class = serializers.RegisterStatisticsSerializer
    queryset = org_models.StudentDiscipline.objects.filter(is_active=True)
    pagination_class = CustomPagination

    def list(self, request, *args, **kwargs):
        profile = request.user.profile

        study_year = request.query_params.get('study_year')
        reg_period = request.query_params.get('reg_period')
        acad_period = request.query_params.get('acad_period')
        faculty = request.query_params.get('faculty')
        speciality = request.query_params.get('speciality')
        edu_prog = request.query_params.get('edu_prog')
        course = request.query_params.get('course')
        group = request.query_params.get('group')
        ordering = request.query_params.getlist('ordering[]')
        queryset = self.queryset.all()
        query = {
            'status_id': student_discipline_status['not_chosen'],
            'study_plan__advisor': profile,
        }

        if acad_period:
            query['acad_period_id'] = acad_period
        elif reg_period:
            acad_period_pks = common_models.CourseAcadPeriodPermission.objects.filter(
                registration_period_id=reg_period,
                # course=course
            ).values('acad_period')
            query['acad_period__in'] = acad_period_pks

        if faculty:
            query['study_plan__faculty_id'] = faculty
        if speciality:
            query['study_plan__speciality_id'] = speciality
        if edu_prog:
            query['study_plan__education_program_id'] = edu_prog
        if group:
            query['study_plan__group_id'] = group
        if study_year:
            query['study_year_id'] = study_year

        if course and study_year:
            study_plan_pks = org_models.StudyYearCourse.objects.filter(
                study_year_id=study_year,
                course=course
            ).values('study_plan')
            query['study_plan__in'] = study_plan_pks
        queryset_uids = queryset.filter(**query).distinct('discipline', 'study_plan__group').values_list('uid', flat=True)
        queryset = queryset.filter(uid__in=queryset_uids).order_by(*ordering)
        distincted_queryset = queryset.values(
            'uid',
            'study_plan__group',
            'study_plan__group_id',
            'study_plan__faculty__name',
            'study_plan__cathedra__name',
            'study_plan__speciality__name',
            'study_plan__group__name',
            'discipline__name',
            'discipline',
            'discipline_id'
        )
        student_discipline_list = []
        page = self.paginate_queryset(distincted_queryset)
        for student_discipline in page:
            query2 = {
                'is_active': True
            }
            query3 = {
                'status__number': 1
            }
            query3.update(query)
            if student_discipline.get('study_plan__group_id'):
                query2['group_id'] = student_discipline.get('study_plan__group_id')
                query3['study_plan__group_id'] = student_discipline.get('study_plan__group_id')

            group_student_count = org_models.StudyPlan.objects.filter(**query2).distinct('student').count()
            if student_discipline.get('discipline_id'):
                query3['discipline_id'] = student_discipline.get('discipline_id')
            not_chosen_student_count = self.queryset.exclude(load_type__load_type2_id='303e48f0-7b1a-431a-aac0-53843479e58e').filter(**query3).distinct('student').count()

            student_discipline['student_count'] = group_student_count
            student_discipline['not_chosen_student_count'] = not_chosen_student_count
            student_discipline['percent_of_non_chosen_student'] = (not_chosen_student_count / group_student_count) * 100

            # d = {
            #     'uid': student_discipline.get('uid'),
            #     'faculty': student_discipline.get('study_plan__faculty__name'),
            #     'cathedra': student_discipline.get('study_plan__cathedra__name'),
            #     'speciality': student_discipline.get('study_plan__speciality__name'),
            #     'group': student_discipline.get('study_plan__group__name'),
            #     'student_count': group_student_count,
            #     'discipline': student_discipline.get('discipline__name'),
            #     'not_chosen_student_count': not_chosen_student_count,
            #     'percent_of_non_chosen_student': (not_chosen_student_count / group_student_count) * 100,
            #     }

            student_discipline_list.append(student_discipline)

        # page = self.paginate_queryset(student_discipline_list)
        if page is not None:
            # serializer = self.serializer_class(page,
            #                                    many=True)
            return self.get_paginated_response(student_discipline_list)


class NotRegisteredStudentListView(generics.ListAPIView):
    """Список незарегистрированных
    study_year(!), reg_period(!), acad_period, faculty, speciality, edu_prog, course, group
    """
    queryset = org_models.StudentDiscipline.objects.filter(is_active=True)
    # serializer_class = serializers.NotRegisteredStudentSerializer
    pagination_class = CustomPagination

    def list(self, request, *args, **kwargs):
        profile = request.user.profile
        study_year = request.query_params.get('study_year')
        reg_period = request.query_params.get('reg_period')
        acad_period = request.query_params.get('acad_period')
        faculty = request.query_params.get('faculty')
        speciality = request.query_params.get('speciality')
        edu_prog = request.query_params.get('edu_prog')
        course = request.query_params.get('course')
        group = request.query_params.get('group')
        ordering = request.query_params.getlist('ordering[]')

        queryset = self.queryset.all()
        # if ordering:
        #     queryset = queryset.order_by(*ordering)

        query = dict()
        queryset = queryset.filter(
            status__number=1,
            study_plan__advisor=profile,
        ).distinct('student')

        if acad_period:
            query['acad_period_id'] = acad_period
        elif reg_period:
            acad_period_pks = common_models.CourseAcadPeriodPermission.objects.filter(
                registration_period_id=reg_period,
                # course=course
            ).values('acad_period')
            query['acad_period__in'] = acad_period_pks

        if faculty:
            query['study_plan__faculty_id'] = faculty
        if speciality:
            query['study_plan__speciality_id'] = speciality
        if edu_prog:
            query['study_plan__education_program_id'] = edu_prog
        if group:
            query['study_plan__group_id'] = group
        if study_year:
            query['study_year_id'] = study_year

        if course and study_year:
            study_plan_pks = org_models.StudyYearCourse.objects.filter(
                study_year_id=study_year,
                course=course
            ).values('study_plan')
            query['study_plan__in'] = study_plan_pks

        distincted_queryset = queryset.filter(**query).distinct(
            'study_plan__faculty',
            'study_plan__cathedra',
            'study_plan__speciality',
            'study_plan__group',
            'discipline'
        )

        if ordering:
            distincted_queryset = self.queryset.filter(uid__in=Subquery(distincted_queryset.values_list('uid', flat=True))).order_by(*ordering)

        student_discipline_list = []
        page = self.paginate_queryset(distincted_queryset)
        for item in page:
            student_name_list = queryset.exclude(load_type__load_type2_id='303e48f0-7b1a-431a-aac0-53843479e58e').filter(
                study_plan__faculty=item.study_plan.faculty,
                study_plan__cathedra=item.study_plan.cathedra,
                study_plan__speciality=item.study_plan.speciality,
                study_plan__group=item.study_plan.group,
                discipline=item.discipline,
            ).annotate(
                fio=Concat(F('student__last_name'),
                           Value(' '), F('student__first_name'),
                           Value(' '), F('student__middle_name'))).values_list('fio', flat=True)

            d = {
                'faculty': item.study_plan.faculty.name,
                'cathedra': item.study_plan.cathedra.name,
                'speciality': item.study_plan.speciality.name,
                'group': item.study_plan.group.name,
                'discipline': item.discipline.name,
                'student': student_name_list,
            }
            student_discipline_list.append(d)

        if page is not None:
            return self.get_paginated_response(student_discipline_list)


class GenerateIupBidExcelView(generics.RetrieveAPIView):
    """Создать запрос на Excel для Заявки на ИУПы
        study_year(!), study_form, faculty, cathedra, edu_prog_group, edu_prog,
        course, group, acad_periods, status, reg_period
    """

    def get(self, request, *args, **kwargs):
        profile = request.user.profile

        study_year = request.query_params.get('study_year')  # Дисциплина студента
        study_form = request.query_params.get('study_form')
        faculty = request.query_params.get('faculty')
        cathedra = request.query_params.get('cathedra')
        edu_prog_group = request.query_params.get('edu_prog_group')
        edu_prog = request.query_params.get('edu_prog')
        course = request.query_params.get('course')  # Дисциплина студента
        group = request.query_params.get('group')
        reg_period = self.request.query_params.get('reg_period')
        status_id = self.request.query_params.get('status')  # В Дисциплине студента
        acad_periods = self.request.query_params.get('acad_periods')

        fields = {
            'study_year': study_year,
            'study_form': study_form,
            'faculty': faculty,
            'cathedra': cathedra,
            'edu_prog_group': edu_prog_group,
            'edu_prog': edu_prog,
            'course': course,
            'group': group,
            'reg_period': reg_period,
            'status_id': status_id,
            'acad_periods': acad_periods,
        }
        token = uuid4()
        fields_json = json.dumps(fields)
        ExcelTask.objects.create(
            doc_type=4,
            profile=profile,
            token=token,
            fields=fields_json,
        )
        return Response(
            {
                'token': str(token)
            },
            status=status.HTTP_200_OK
        )


def make_iup_bid_excel(task):
    """Генерировать Excel для Заявки на ИУПы
        study_year(!), study_form, faculty, cathedra, edu_prog_group, edu_prog,
        course, group, acad_periods, status, reg_period
    """

    profile = task.profile
    fields_json = task.fields
    fields = json.loads(fields_json)

    study_year = fields.get('study_year')  # Дисциплина студента
    study_form = fields.get('study_form')
    faculty = fields.get('faculty')
    cathedra = fields.get('cathedra')
    edu_prog_group = fields.get('edu_prog_group')
    edu_prog = fields.get('edu_prog')
    course = fields.get('course')  # Дисциплина студента
    group = fields.get('group')
    reg_period = fields.get('reg_period')
    status_id = fields.get('status_id')  # В Дисциплине студента
    acad_periods = fields.get('acad_periods')

    wb = load_workbook('advisors/excel/template.xlsx')
    ws = wb.active
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))
    font = Font(
        name='Times New Roman',
        size=11
    )
    font_bold = Font(
        name='Times New Roman',
        size=11,
        bold=True,
    )
    alignment = Alignment(
        vertical="center",
        horizontal="center"
    )
    wrap_alignment = Alignment(
        wrapText=True,
        vertical="center",
        horizontal="center",
    )
    wrap_left_alignment = Alignment(
        wrapText=True,
        vertical="center",
        horizontal="left",
    )

    queryset = org_models.StudyPlan.objects.filter(
        is_active=True,
        advisor=profile,
    )

    queryset = queryset.exclude(student__status_id=STUDENT_STATUSES['expelled'])

    if reg_period:
        reg_period_obj = common_models.RegistrationPeriod.objects.get(pk=reg_period)
        ws['B4'] = reg_period_obj.name
        ws['B4'].font = font
        ws['B4'].alignment = wrap_left_alignment

    if course:
        ws['B5'] = course
    else:
        ws['B5'] = 'Все'
    ws['B5'].font = font

    if status_id:
        status_obj = org_models.StudentDisciplineStatus.objects.get(number=status_id)
        study_plan_pks_from_sd = org_models.StudentDiscipline.objects.filter(
            study_year_id=study_year,
            status=status_obj,
        ).values('study_plan')
        queryset = queryset.filter(pk__in=study_plan_pks_from_sd)

        ws['B6'] = status_obj.name
    else:
        ws['B6'] = 'Все'
    ws['B6'].font = font

    if study_form:
        study_form_obj = org_models.StudyForm.objects.get(pk=study_form)
        ws['B7'] = study_form_obj.name
        ws['B7'].alignment = wrap_left_alignment
        queryset = queryset.filter(study_form_id=study_form)
    else:
        ws['B7'] = 'Все'
    ws['B7'].font = font

    if faculty:
        faculty_obj = org_models.Faculty.objects.get(pk=faculty)
        queryset = queryset.filter(faculty_id=faculty)
        ws['B8'] = faculty_obj.name
        ws['B8'].alignment = wrap_left_alignment
    else:
        ws['B8'] = 'Все'
    ws['B8'].font = font

    if cathedra:
        cathedra_obj = org_models.Cathedra.objects.get(pk=cathedra)
        queryset = queryset.filter(cathedra_id=cathedra)
        ws['B9'] = cathedra_obj.name
        ws['B9'].alignment = wrap_left_alignment
    else:
        ws['B9'] = 'Все'
    ws['B9'].font = font

    if edu_prog_group:
        queryset = queryset.filter(education_program__group_id=edu_prog_group)
        edu_prog_group_obj = org_models.EducationProgramGroup.objects.get(pk=edu_prog_group)
        ws['B10'] = edu_prog_group_obj.name
        ws['B10'].alignment = wrap_left_alignment
    else:
        ws['B10'] = 'Все'
    ws['B10'].font = font

    if edu_prog:
        queryset = queryset.filter(education_program_id=edu_prog)
        edu_prog_obj = org_models.EducationProgram.objects.get(pk=edu_prog)
        ws['B11'] = edu_prog_obj.name
        ws['B11'].alignment = wrap_left_alignment
    else:
        ws['B11'] = 'Все'
    ws['B11'].font = font

    if group:
        queryset = queryset.filter(group_id=group)
        group_obj = org_models.Group.objects.get(pk=group)
        ws['B12'] = group_obj.name
        ws['B12'].alignment = wrap_left_alignment
    else:
        ws['B12'] = 'Все'
    ws['B12'].font = font

    if study_year:
        study_year_obj = org_models.StudyPeriod.objects.get(pk=study_year)
        queryset = queryset.filter(study_period__end__gt=study_year_obj.start)
        ws['B3'] = '{} - {}'.format(study_year_obj.start,
                                    study_year_obj.end)
        ws['B3'].font = font

    if course and study_year:
        study_plan_pks = org_models.StudyYearCourse.objects.filter(
            study_year_id=study_year,
            course=course
        ).values('study_plan')
        queryset = queryset.filter(pk__in=study_plan_pks)

    now = datetime.now()
    ws['B13'] = now.strftime("%d:%m:%Y, %H:%M:%S")
    ws['B13'].font = font

    for i, study_plan in enumerate(queryset):
        row_num = str(15 + i)

        a = 'A' + row_num
        ws[a] = i + 1
        ws[a].font = font
        ws[a].alignment = alignment
        ws[a].border = border

        b = 'B' + row_num
        ws[b] = study_plan.education_program.group.code if study_plan.education_program.group else ' '  # study_plan.education_program.group.code
        ws[b].font = font
        ws[b].alignment = wrap_alignment
        ws[b].border = border

        c = 'C' + row_num
        ws[c] = study_plan.education_program.name
        ws[c].font = font
        ws[c].alignment = wrap_alignment
        ws[c].border = border

        d = 'D' + row_num
        ws[d] = study_plan.group.name if study_plan.group else ''
        ws[d].font = font
        ws[d].alignment = wrap_alignment
        ws[d].border = border

        e = 'E' + row_num
        ws[e] = study_plan.student.full_name
        ws[e].font = font
        ws[e].alignment = wrap_alignment
        ws[e].border = border

        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J',
                   'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']

        current_col_num = 5
        sum_credit = 0
        old_status = 0
        mark_text = ''

        if acad_periods and len(acad_periods) > 0:
            acad_period_list = acad_periods.split(',')

            for i, acad_period in enumerate(acad_period_list):
                acad_period_obj = org_models.AcadPeriod.objects.get(pk=acad_period)
                head = '{} триместр Перечень дисциплин, ' \
                       'на которые проведена  регистрация ' \
                       'с указанием кредитов'.format(acad_period_obj.number)

                student_disciplines = org_models.StudentDiscipline.objects.filter(
                    is_active=True,
                    study_plan=study_plan,
                    study_year_id=study_year,
                    acad_period_id=acad_period,
                ).distinct('discipline')

                student_discipline_list = list(student_disciplines)
                credit_list = [i.credit for i in student_discipline_list]
                total_credit = sum(credit_list)
                sum_credit += total_credit

                text = 'Кредиты: {}\n'.format(total_credit)

                for stud_discipline in student_disciplines:
                    text += '{} ({}), '.format(stud_discipline.discipline.name,
                                               stud_discipline.credit)

                head_cell = columns[current_col_num] + '14'
                ws[head_cell] = head
                ws[head_cell].font = font_bold
                ws[head_cell].border = border
                # ws[head_cell].alignment = alignment

                text_cell = columns[current_col_num] + row_num
                ws[text_cell] = text
                ws[text_cell].font = font
                ws[text_cell].border = border
                ws[text_cell].alignment = wrap_alignment

                ws.row_dimensions[int(row_num)].height = 70

                # if len(text) > 200:
                #     ws.row_dimensions[row_num].height = 50  # Test

                current_col_num += 1

                checks = models.AdvisorCheck.objects.filter(
                    study_plan=study_plan,
                    acad_period_id=acad_period,
                )
                if checks.exists():
                    old_status = checks.latest('id').status
                else:
                    old_status = 0

                if old_status == 3:
                    mark = '{} - Отклонено\n'.format(acad_period_obj.repr_name)
                elif old_status == 4:
                    mark = '{} - Утверждено\n'.format(acad_period_obj.repr_name)
                elif old_status == 5:
                    mark = '{} - Изменено\n'.format(acad_period_obj.repr_name)
                else:
                    mark = '{} - Не определено\n'.format(acad_period_obj.repr_name)

                mark_text += mark

        gos_attestation_label = columns[current_col_num] + '14'
        ws[gos_attestation_label] = 'Государственная аттестация'
        ws[gos_attestation_label].font = font_bold
        ws[gos_attestation_label].alignment = alignment
        ws[gos_attestation_label].border = border

        gos_attestation = columns[current_col_num] + row_num
        ws[gos_attestation] = 'не выбрана'
        ws[gos_attestation].font = font
        ws[gos_attestation].alignment = alignment
        ws[gos_attestation].border = border

        current_col_num += 1
        sum_credit_label = columns[current_col_num] + '14'
        ws[sum_credit_label] = 'Общее количество кредитов'
        ws[sum_credit_label].font = font_bold
        ws[sum_credit_label].alignment = alignment
        ws[sum_credit_label].border = border

        sum_credit_cell = columns[current_col_num] + row_num
        ws[sum_credit_cell] = sum_credit
        ws[sum_credit_cell].font = font
        ws[sum_credit_cell].alignment = alignment
        ws[sum_credit_cell].border = border

        current_col_num += 1
        advisor_mark_label = columns[current_col_num] + '14'
        ws[advisor_mark_label] = 'Отметка эдвайзера'
        ws[advisor_mark_label].font = font_bold
        ws[advisor_mark_label].alignment = alignment
        ws[advisor_mark_label].border = border

        mark_cell = columns[current_col_num] + row_num
        ws[mark_cell] = mark_text
        ws[mark_cell].font = font
        ws[mark_cell].border = border
        ws[mark_cell].alignment = wrap_alignment

    file_name = 'temp_files/zayavki{}.xlsx'.format(str(uuid4()))
    wb.save(file_name)
    task.file_path = file_name
    task.save()


class GenerateIupExcelView(generics.RetrieveAPIView):
    """
    Создать запрос на Excel ИУПы обучающихся (2 стр, Утвержденные)
    study_year(!), edu_prog(!), student(!), reg_period(!) faculty, speciality, group,
    """

    def get(self, request, *args, **kwargs):
        profile = request.user.profile

        study_year = request.query_params.get('study_year')
        edu_prog = request.query_params.get('edu_prog')
        student = request.query_params.get('student')
        reg_period = self.request.query_params.get('reg_period')

        faculty = request.query_params.get('faculty')
        speciality = request.query_params.get('speciality')
        group = request.query_params.get('group')

        fields = {
            'study_year': study_year,
            'faculty': faculty,
            'edu_prog': edu_prog,
            'group': group,
            'reg_period': reg_period,
            'speciality': speciality,
            'student': student,
        }
        token = uuid4()
        fields_json = json.dumps(fields)
        ExcelTask.objects.create(
            doc_type=5,
            profile=profile,
            token=token,
            fields=fields_json,
        )
        return Response(
            {
                'token': str(token)
            },
            status=status.HTTP_200_OK
        )


def make_iup_excel(task):
    """
    Excel ИУПы обучающихся (2 стр, Утвержденные)
    study_year(!), edu_prog(!), student(!), reg_period(!) faculty, speciality, group,
    """

    profile = task.profile
    fields_json = task.fields
    fields = json.loads(fields_json)

    study_year = fields.get('study_year')
    edu_prog = fields.get('edu_prog')
    student = fields.get('student')
    reg_period = fields.get('reg_period')

    faculty = fields.get('faculty')
    speciality = fields.get('speciality')
    group = fields.get('group')

    wb = load_workbook('advisors/excel/template2.xlsx')
    ws = wb.active

    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))
    bottom_border = Border(bottom=Side(border_style='thin', color='000000'))
    font = Font(
        name='Times New Roman',
        size=11
    )
    font_bold = Font(
        name='Times New Roman',
        bold=True,
        size=11
    )
    font_large = Font(
        name='Times New Roman',
        bold=True,
        size=12
    )
    left_alignment = Alignment(
        horizontal="left"
    )
    filters = {}

    if faculty:
        filters['faculty_id'] = faculty

    if speciality:
        filters['speciality_id'] = speciality

    if group:
        filters['group_id'] = group

    try:
        study_year_obj = org_models.StudyPeriod.objects.get(pk=study_year)
        study_plan = org_models.StudyPlan.objects.get(
            advisor=profile,
            study_period__end__gt=study_year_obj.start,
            education_program_id=edu_prog,
            student_id=student,
            is_active=True,
            **filters,
        )
    except org_models.StudyPlan.DoesNotExist:
        return Response(
            {
                'message': 0  # Учебный план не найден
            },
            status=status.HTTP_404_NOT_FOUND
        )
    try:
        decan_name_cell = 'E10'
        p = study_plan.faculty.dekan
        ws[decan_name_cell] = "{} {} {}".format(p.last_name, p.first_name, p.middle_name)
        ws[decan_name_cell].font = font
    except:
        pass

    student_name_cell = 'C17'
    ws[student_name_cell] = study_plan.student.full_name
    ws[student_name_cell].font = font

    acad_degree_cell = 'C18'
    ws[acad_degree_cell] = study_plan.preparation_level.name

    speciality_cell = 'C19'
    ws[speciality_cell] = '{} ({})'.format(study_plan.speciality.name,
                                           study_plan.speciality.code)

    study_form_cell = 'C20'
    max_course = \
        org_models.StudyYearCourse.objects.filter(study_plan=study_plan).aggregate(max_course=Max('course'))[
            'max_course']
    ws[study_form_cell] = '{}, {} года'.format(study_plan.study_form.name, max_course)

    current_course_cell = 'C21'
    ws[current_course_cell] = study_plan.current_course
    ws[current_course_cell].alignment = left_alignment

    lang_cell = 'C22'
    ws[lang_cell] = study_plan.group.language.name

    current_study_year_cell = 'C23'
    study_year_dict = get_current_study_year()
    ws[current_study_year_cell] = '{}-{}'.format(study_year_dict['start'],
                                                 study_year_dict['end'])

    course = study_plan.get_course(study_year_obj)
    ws['D27'] = '{} Курс обучения  {} учебный год'.format(course,
                                                          str(study_year_obj))
    ws['D27'].font = font_large

    acad_period_pks_from_sd = org_models.StudentDiscipline.objects.filter(
        study_year_id=study_year,
        study_plan=study_plan,
    ).distinct('acad_period').values('acad_period')

    acad_periods = org_models.AcadPeriod.objects.filter(
        pk__in=acad_period_pks_from_sd,
    )

    acad_period_pks_from_rule = common_models.CourseAcadPeriodPermission.objects.filter(
        registration_period_id=reg_period,
        course=course,
    ).values('acad_period')
    acad_periods = acad_periods.filter(pk__in=acad_period_pks_from_rule)

    row_num = 28
    sd_num = 1
    total_credit_in_course = 0

    for acad_period in acad_periods:
        if StudentDisciplineInfo.objects.filter(
                study_plan=study_plan,
                acad_period=acad_period,
                status_id=student_discipline_info_status['confirmed'],
        ).exists():
            student_disciplines = org_models.StudentDiscipline.objects.filter(
                study_year_id=study_year,
                study_plan=study_plan,
                acad_period=acad_period,
            ).distinct('discipline')

            ws['D' + str(row_num)] = acad_period.repr_name
            ws['D' + str(row_num)].font = font_bold
            row_num += 1

            total_credit_in_acad_period = 0

            for sd in student_disciplines:
                num_cell = 'A' + str(row_num)
                ws[num_cell] = sd_num
                ws[num_cell].border = border
                ws[num_cell].font = font

                component_cell = 'B' + str(row_num)
                if sd.component:
                    ws[component_cell] = sd.component.short_name
                elif sd.cycle:
                    ws[component_cell] = sd.cycle.short_name
                else:
                    ws[component_cell] = ''

                ws[component_cell].border = border
                ws[component_cell].font = font

                discipline_code_cell = 'C' + str(row_num)
                ws[discipline_code_cell] = sd.discipline_code
                ws[discipline_code_cell].border = border
                ws[discipline_code_cell].font = font

                discipline_name_cell = 'D' + str(row_num)
                ws[discipline_name_cell] = sd.discipline.name
                ws[discipline_name_cell].border = border
                ws[discipline_name_cell].font = font

                credit_cell = 'E' + str(row_num)
                ws[credit_cell] = sd.credit
                ws[credit_cell].border = border
                ws[credit_cell].font = font

                total_credit_in_acad_period += sd.credit

                row_num += 1
                sd_num += 1

            total_credit_in_course += total_credit_in_acad_period
            ws['B' + str(row_num)] = 'Общее количество кредитов'
            ws['B' + str(row_num)].font = font
            ws['E' + str(row_num)] = total_credit_in_acad_period
            ws['E' + str(row_num)].font = font

            ws['B' + str(row_num)].border = border
            ws['A' + str(row_num)].border = border
            ws['C' + str(row_num)].border = bottom_border
            ws['D' + str(row_num)].border = bottom_border
            ws['E' + str(row_num)].border = border

            row_num += 1

    ws['B' + str(row_num)] = 'Общее количество кредитов за курс'
    ws['B' + str(row_num)].font = font
    ws['E' + str(row_num)] = total_credit_in_course
    ws['E' + str(row_num)].font = font

    ws['B' + str(row_num)].border = border
    ws['A' + str(row_num)].border = border
    ws['C' + str(row_num)].border = bottom_border
    ws['D' + str(row_num)].border = bottom_border
    ws['E' + str(row_num)].border = border

    # row_num += 2
    # ws['A' + str(row_num)] = 'Регистратор'
    # ws['A' + str(row_num)].font = font

    row_num += 2
    p = study_plan.advisor
    ws['A' + str(row_num)] = "Эдвайзер {} {} {}".format(p.last_name, p.first_name, p.middle_name)
    ws['A' + str(row_num)].font = font

    row_num += 2
    p = study_plan.student
    ws['A' + str(row_num)] = "Обучающийся {} {} {}".format(p.last_name, p.first_name, p.middle_name)
    ws['A' + str(row_num)].font = font

    file_name = 'temp_files/iupi{}.xlsx'.format(str(uuid4()))
    wb.save(file_name)
    task.file_path = file_name
    task.save()


class CopyStudyPlansListView(generics.ListAPIView):
    """
    Получение учебных планов,
    study_year(!), study_form, faculty, cathedra, edu_prog_group, edu_prog, course, group, status
    """
    queryset = org_models.StudyPlan.objects.filter(is_active=True).order_by('student__last_name')
    serializer_class = serializers.StudyPlanSerializer
    pagination_class = AdvisorBidPagination  # CustomPagination

    def list(self, request, *args, **kwargs):
        study_year = request.query_params.get('study_year')
        study_form = request.query_params.get('study_form')
        faculty = request.query_params.get('faculty')
        cathedra = request.query_params.get('cathedra')
        edu_prog_group = request.query_params.get('edu_prog_group')
        edu_prog = request.query_params.get('edu_prog')
        course = request.query_params.get('course')
        group = request.query_params.get('group')

        status_id = request.query_params.get('status')
        acad_periods = request.query_params.get('acad_periods')
        #reg_period = self.request.query_params.get('reg_period')

        queryset = self.queryset.filter(advisor=self.request.user.profile)
        queryset = queryset.exclude(student__status_id=STUDENT_STATUSES['expelled'])

        sd = org_models.StudentDiscipline.objects.filter(study_year_id=study_year)

        if status_id:
            status_obj = org_models.StudentDisciplineStatus.objects.get(number=status_id)

            study_plan_pks_from_sd = org_models.StudentDiscipline.objects.filter(
                study_year_id=study_year,
                status=status_obj,
            ).values('study_plan')

            sd = sd.filter(status=status_obj)
            queryset = queryset.filter(pk__in=study_plan_pks_from_sd)
        lookup = Q()
        sd_lookup = Q()
        if study_form:
            lookup = lookup & Q(study_form_id=study_form)
            sd_lookup = sd_lookup & Q(study_plan__study_form_id=study_form)
            # queryset = queryset.filter(study_form_id=study_form)
            # sd = sd.filter(study_plan__study_form_id=study_form)
        if faculty:
            lookup = lookup & Q(faculty_id=faculty)
            sd_lookup = sd_lookup & Q(study_plan__faculty_id=faculty)
            # queryset = queryset.filter(faculty_id=faculty)
            # sd = sd.filter(study_plan__faculty_id=faculty)
        if cathedra:
            lookup = lookup & Q(cathedra_id=cathedra)
            sd_lookup = sd_lookup & Q(study_plan__cathedra_id=cathedra)
        #     queryset = queryset.filter(cathedra_id=cathedra)
        #     sd = sd.filter(study_plan__cathedra_id=cathedra)
        if edu_prog:
            lookup = lookup & Q(education_program_id=edu_prog)
            sd_lookup = sd_lookup & Q(study_plan__education_program_id=edu_prog)
            # queryset = queryset.filter(education_program_id=edu_prog)
            # sd = sd.filter(study_plan__education_program_id=edu_prog)
        if edu_prog_group:
            lookup = lookup & Q(education_program__group_id=edu_prog_group)
            sd_lookup = sd_lookup & Q(study_plan__education_program__group_id=edu_prog_group)
            # queryset = queryset.filter(education_program__group_id=edu_prog_group)
            # sd = sd.filter(study_plan__education_program__group_id=edu_prog_group)
        if group:
            lookup = lookup & Q(group_id=group)
            sd_lookup = sd_lookup & Q(study_plan__group_id=group)
            # queryset = queryset.filter(group_id=group)
            # sd = sd.filter(study_plan__group_id=group)
        if study_year:
            study_year_obj = org_models.StudyPeriod.objects.get(pk=study_year)
            queryset = queryset.filter(study_period__end__gt=study_year_obj.start)
        queryset = queryset.filter(lookup)
        sd = sd.filter(sd_lookup)
        if course and study_year:
            study_plan_pks = org_models.StudyYearCourse.objects.filter(
                study_year_id=study_year,
                course=course
            ).values('study_plan')
            queryset = queryset.filter(pk__in=study_plan_pks)

        reg_period = common_models.RegistrationPeriod.objects.get(study_year=study_year)

        if course:
            acad_period_pks = common_models.CourseAcadPeriodPermission.objects.filter(
                registration_period_id=reg_period,
                course=course
            ).values_list('acad_period', flat=True)
        else:
            acad_period_pks = common_models.CourseAcadPeriodPermission.objects.filter(
                registration_period_id=reg_period,
            ).values_list('acad_period', flat=True)

        acad_periods_repr = org_models.AcadPeriod.objects.filter(is_active=True,
                                                                 pk__in=acad_period_pks,)
        acad_period_pks_from_sd = sd.values('acad_period')
        acad_periods_repr = acad_periods_repr.filter(pk__in=acad_period_pks_from_sd)

        sd = sd.filter(acad_period__in=acad_periods_repr).values_list('student', flat=True).distinct()

        queryset = queryset.filter(student__in=sd)

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.serializer_class(page,
                                               many=True,
                                               context={'study_year': study_year,
                                                        'status_id': status_id,
                                                        'acad_periods': acad_periods})
            return self.get_paginated_response(serializer.data)


class DeactivateDiscipline(generics.UpdateAPIView):
    """
    Деактивировать дисциплины
    """
    permission_classes = (
        IsAuthenticated,
        # adv_permission.StudentDisciplinePermission,
    )
    queryset = org_models.StudentDiscipline.objects.all()
    serializer_class = serializers.DeactivateDisciplineSerializer


class ActivateDiscipline(generics.UpdateAPIView):
    """
    Деактивировать дисциплины
    """
    permission_classes = (
        IsAuthenticated,
        # adv_permission.StudentDisciplinePermission,
    )
    queryset = org_models.StudentDiscipline.objects.all()
    serializer_class = serializers.ActivateDisciplineSerializer


class StudentProfilesList(generics.ListAPIView):
    queryset = org_models.StudyPlan.objects.filter(is_active=True).order_by('student__last_name')
    serializer_class = serializers.StudentProfilesListSerializer
    pagination_class = AdvisorBidPagination  # CustomPagination

    def list(self, request, *args, **kwargs):
        queryset = self.queryset.filter(advisor=self.request.user.profile)

        entry_year = request.query_params.get('entry_year')
        study_form = request.query_params.get('study_form')
        preparation_level = request.query_params.get('preparation_level')
        faculty = request.query_params.get('faculty')
        cathedra = request.query_params.get('cathedra')
        speciality = request.query_params.get('speciality')
        edu_prog_group = request.query_params.get('edu_prog_group')
        edu_prog = request.query_params.get('edu_prog')
        course = request.query_params.get('course')
        group = request.query_params.get('group')
        student_status = request.query_params.get('student_status')
        gender = request.query_params.get('gender')
        citizenship = request.query_params.get('citizenship')

        lookup = Q()

        if entry_year:
            lookup = lookup & Q(entry_date__year=entry_year)
        if study_form:
            lookup = lookup & Q(study_form_id=study_form)
        if faculty:
            lookup = lookup & Q(faculty_id=faculty)
        if cathedra:
            lookup = lookup & Q(cathedra_id=cathedra)
        if edu_prog:
            lookup = lookup & Q(education_program_id=edu_prog)
        if edu_prog_group:
            lookup = lookup & Q(education_program__group_id=edu_prog_group)
        if group:
            lookup = lookup & Q(group_id=group)
        if preparation_level:
            lookup = lookup & Q(preparation_level_id=preparation_level)
        if speciality:
            lookup = lookup & Q(speciality_id=speciality)
        if course:
            study_plan_pks = org_models.StudyYearCourse.objects.filter(course=course).values('study_plan')
            lookup = lookup & Q(pk__in=study_plan_pks)
        if student_status:
            lookup = lookup & Q(student__status_id=student_status)
        if gender:
            lookup = lookup & Q(student__gender=gender)
        if citizenship:
            lookup = lookup & Q(student__citizenship=citizenship)

        queryset = queryset.filter(lookup)

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.serializer_class(page,
                                               many=True, )
            return self.get_paginated_response(serializer.data)


class StudentsByDisplinesIDListView(generics.ListAPIView):
    queryset = org_models.StudentDiscipline.objects.filter(is_active=True)
    serializer_class = serializers.StudentsByDisciplineIDSerializer
    pagination_class = CustomPagination

    def list(self, request, *args, **kwargs):
        queryset = self.queryset.filter(teacher=self.request.user.profile).exclude(
            student__status_id=STUDENT_STATUSES['expelled'])

        discipline_id = request.query_params.get('discipline_id')
        load_type_id = request.query_params.get('load_type_id')
        language_id = request.query_params.get('language_id')
        hours = request.query_params.get('hours')

        study_year = request.query_params.get('study_year')
        faculty = request.query_params.get('faculty')
        edu_prog = request.query_params.get('edu_prog')
        course = request.query_params.get('course')
        group = request.query_params.get('group')

        acad_period = request.query_params.get('acad_period')
        reg_period = request.query_params.get('reg_period')

        lookup = Q(status_id=student_discipline_status['confirmed'],
                   discipline_id=discipline_id,
                   load_type_id=load_type_id,
                   language_id=language_id,
                   hours=hours)

        queryset = queryset.filter(lookup)

        if acad_period:
            lookup = lookup & Q(acad_period_id=acad_period)

        if reg_period:
            query_temp = common_models.CourseAcadPeriodPermission.objects.filter(
                registration_period_id=reg_period
            ).values('acad_period')
            lookup = lookup & Q(acad_period__in=query_temp)

        if faculty:
            lookup = lookup & Q(study_plan__faculty_id=faculty)

        if edu_prog:
            lookup = lookup & Q(study_plan__education_program_id=edu_prog)

        if group:
            lookup = lookup & Q(study_plan__group_id=group)

        if study_year:
            lookup = lookup & Q(study_year_id=study_year)

        if course and study_year:
            query_temp = org_models.StudyYearCourse.objects.filter(
                study_year_id=study_year,
                course=course,
            ).values('study_plan')
            lookup = lookup & Q(study_year__in=query_temp)

        queryset = queryset.filter(lookup)

        page = self.paginate_queryset(queryset)

        if page is not None:
            serializer = self.serializer_class(page, many=True)
            return self.get_paginated_response(serializer.data)


class ThesisTopic(APIView):

    def get(self, request, format=None):
        result = {'status': False}
        if request.GET.get('uid_1c'):
            theme = models.ThemesOfTheses.objects.filter(
                    uid_1c=request.GET.get('uid_1c'),
                    student__isnull=True
                )
            result['themes_count'] = theme.count()
            result['themes'] = serializers.ThemesOfThesesSerializer(
                theme, many=True).data
            try:
                result['check_themes'] = serializers.ThemesOfThesesSerializer(
                    models.ThemesOfTheses.objects.get(
                        uid_1c=request.GET.get('uid_1c'),
                        student=request.user.profile
                    )).data
            except:
                pass
        else:
            disciplinecredits = org_models.DisciplineCredit.objects.filter(
                chosen_control_forms__is_diploma=True,
                student=request.user.profile,
                is_active=True
             )
            if disciplinecredits.count() > 0:
                result['status'] = True
        return Response(result, status=status.HTTP_200_OK)

    def post(self, request, format=None):
        obj = {"send": "Not theme_uid"}
        if request.data.get('theme_uid'):
            try:
                tem = models.ThemesOfTheses.objects.get(uid=request.data.get('theme_uid'))
                tem.student = request.user.profile
                tem.save()
                obj['theme'] = serializers.ThemesOfThesesSerializer(
                    models.ThemesOfTheses.objects.get(uid=request.data.get('theme_uid'))
                ).data
                obj['send'] = "Ok"
                return Response(obj, status=status.HTTP_200_OK)
            except:
                obj['send'] = 'Not check'
                return Response(obj, status=status.HTTP_400_BAD_REQUEST)
        return Response(obj, status=status.HTTP_400_BAD_REQUEST)


class EntryYearListView(generics.ListAPIView):
    queryset = org_models.StudyPlan.objects.filter(is_active=True).order_by('entry_date__year')
    serializer_class = serializers.EntryYearSerializer

    def get_queryset(self):
        queryset = self.queryset.filter(advisor=self.request.user.profile).distinct('entry_date__year')
        return queryset


class PreparationsLevelListView(generics.ListAPIView):
    queryset = org_models.PreparationLevel.objects.filter(is_active=True).order_by('name')
    serializer_class = serializers.PrepartionLevelListSerializer

    def get_queryset(self):
        study_plans = org_models.StudyPlan.objects.filter(advisor=self.request.user.profile,
                                                          is_active=True).values('preparation_level')
        queryset = self.queryset.filter(pk__in=study_plans)
        return queryset

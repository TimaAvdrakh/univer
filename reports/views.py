from rest_framework import generics
from rest_framework import status
from rest_framework.response import Response
from uuid import uuid4
from django.shortcuts import HttpResponse
from organizations import models as org_models
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl import Workbook, load_workbook
from portal.curr_settings import student_discipline_info_status, student_discipline_status, STUDENT_STATUSES
from common import models as common_models
from datetime import datetime
from . import models
import json
from cron_app.models import ExcelTask
from rest_framework.permissions import IsAuthenticated
from . import permissions


class RegisterResultExcelView(generics.RetrieveAPIView):
    """Cоздать запрос на Результат регистрации Excel
        study_year(!), reg_period(!), acad_period, faculty, speciality, edu_prog, course, group
    """
    queryset = org_models.StudentDiscipline.objects.filter(is_active=True)

    def get(self, request, *args, **kwargs):
        profile = request.user.profile

        study_year = request.query_params.get('study_year')
        reg_period = request.query_params.get('reg_period')
        acad_period = self.request.query_params.get('acad_period')
        faculty = request.query_params.get('faculty')
        speciality = request.query_params.get('speciality')
        edu_prog = request.query_params.get('edu_prog')
        course = request.query_params.get('course')
        group = request.query_params.get('group')

        fields = {
            'study_year':  study_year,
            'reg_period': reg_period,
            'acad_period': acad_period,
            'faculty': faculty,
            'speciality': speciality,
            'edu_prog': edu_prog,
            'course': course,
            'group': group,
        }
        token = uuid4()
        fields_json = json.dumps(fields)
        ExcelTask.objects.create(
            doc_type=1,
            profile=profile,
            token=token,
            fields=fields_json,
        )
        return Response(
            {
                'token': str(token)
            },
            status=status.HTTP_200_OK
        )


def make_register_result_rxcel(task):
    """Создать Результат регистрации Excel
        study_year(!), reg_period(!), acad_period, faculty, speciality, edu_prog, course, group
    """
    queryset = org_models.StudentDiscipline.objects.filter(is_active=True)
    queryset = queryset.exclude(student__status_id=STUDENT_STATUSES['expelled'])

    profile = task.profile
    fields_json = task.fields
    fields = json.loads(fields_json)

    study_year = fields.get('study_year')
    reg_period = fields.get('reg_period')
    acad_period = fields.get('acad_period')
    faculty = fields.get('faculty')
    speciality = fields.get('speciality')
    edu_prog = fields.get('edu_prog')
    course = fields.get('course')
    group = fields.get('group')

    queryset = queryset.filter(
        status_id=student_discipline_status['confirmed'],
        study_plan__advisor=profile,
    )

    wb = load_workbook('advisors/excel/register_result.xlsx')
    ws = wb.active
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))
    font = Font(
        name='Times New Roman',
        size=11
    )
    font_small = Font(
        name='Times New Roman',
        size=8,
    )
    alignment = Alignment(
        vertical="center",
        horizontal="center"
    )

    wrap_left_alignment = Alignment(
        wrapText=True,
        vertical="center",
        horizontal="left",
    )

    if reg_period:
        reg_period_obj = common_models.RegistrationPeriod.objects.get(pk=reg_period)
        ws['B6'] = reg_period_obj.name
        ws['B6'].font = font_small
        ws['B6'].alignment = wrap_left_alignment

    if acad_period:
        queryset = queryset.filter(acad_period_id=acad_period)
        acad_period_obj = org_models.AcadPeriod.objects.get(pk=acad_period)
        ws['B7'] = acad_period_obj.repr_name
        ws['B7'].alignment = wrap_left_alignment
    else:
        ws['B7'] = 'Все'
        acad_period_pks = common_models.CourseAcadPeriodPermission.objects.filter(
            registration_period_id=reg_period,
            # course=course
        ).values('acad_period')
        queryset = queryset.filter(acad_period__in=acad_period_pks)

    ws['B7'].font = font_small

    if faculty:
        queryset = queryset.filter(study_plan__faculty_id=faculty)
        faculty_obj = org_models.Faculty.objects.get(pk=faculty)
        ws['B8'] = faculty_obj.name
        ws['B8'].alignment = wrap_left_alignment
    else:
        ws['B8'] = 'Все'
    ws['B8'].font = font_small

    if speciality:
        queryset = queryset.filter(study_plan__speciality_id=speciality)
        speciality_obj = org_models.Speciality.objects.get(pk=speciality)
        ws['B9'] = speciality_obj.name
        ws['B9'].alignment = wrap_left_alignment
    else:
        ws['B9'] = 'Все'
    ws['B9'].font = font_small

    if edu_prog:
        queryset = queryset.filter(study_plan__education_program_id=edu_prog)
        edu_prog_obj = org_models.EducationProgram.objects.get(pk=edu_prog)
        ws['B10'] = edu_prog_obj.name
        ws['B10'].alignment = wrap_left_alignment
    else:
        ws['B10'] = 'Все'
    ws['B10'].font = font_small

    if course:
        ws['B11'] = course
    else:
        ws['B11'] = 'Все'
    ws['B11'].font = font_small

    if group:
        queryset = queryset.filter(study_plan__group_id=group)
        group_obj = org_models.Group.objects.get(pk=group)
        ws['B12'] = group_obj.name
        ws['B12'].alignment = wrap_left_alignment
    else:
        ws['B12'] = 'Все'
    ws['B12'].font = font_small

    now = datetime.now()
    ws['B13'] = now.strftime("%d:%m:%Y, %H:%M:%S")
    ws['B13'].font = font_small

    if study_year:
        study_year_obj = org_models.StudyPeriod.objects.get(pk=study_year)
        ws['B5'] = '{} - {}'.format(study_year_obj.start,
                                    study_year_obj.end)
        ws['B5'].font = font_small
        ws['B5'].alignment = wrap_left_alignment

        queryset = queryset.filter(study_year_id=study_year)

    if course and study_year:
        study_plan_pks = org_models.StudyYearCourse.objects.filter(
            study_year_id=study_year,
            course=course
        ).values('study_plan')
        queryset = queryset.filter(study_plan__in=study_plan_pks)

    distincted_queryset = queryset.distinct('discipline', 'load_type', 'hours', 'language', 'teacher')

    student_discipline_list = []
    for item in distincted_queryset:
        student_count = queryset.filter(
            discipline=item.discipline,
            load_type=item.load_type,
            language=item.language,
            teacher=item.teacher,
            hours=item.hours,
        ).distinct('student').count()
        d = {
            'discipline': item.discipline.name,
            'load_type': item.load_type.name,
            'hours': item.hours,
            'language': item.language.name,
            'teacher': item.teacher.full_name if item.teacher else '',
            'student_count': student_count
        }
        student_discipline_list.append(d)

    for i, sd in enumerate(student_discipline_list):
        row_num = str(15 + i)
        a = 'A' + row_num
        ws[a] = sd['discipline']
        ws[a].font = font
        ws[a].border = border
        ws[a].alignment = wrap_left_alignment

        b = 'B' + row_num
        ws[b] = sd['load_type']
        ws[b].font = font
        ws[b].border = border
        ws[b].alignment = wrap_left_alignment

        c = 'C' + row_num
        ws[c] = sd['hours']
        ws[c].font = font
        ws[c].border = border
        ws[c].alignment = wrap_left_alignment

        d = 'D' + row_num
        ws[d] = sd['language']
        ws[d].font = font
        ws[d].border = border
        ws[d].alignment = wrap_left_alignment

        e = 'E' + row_num
        ws[e] = sd['teacher']
        ws[e].font = font
        ws[e].border = border
        ws[e].alignment = wrap_left_alignment

        f = 'F' + row_num
        ws[f] = sd['student_count']
        ws[f].font = font
        ws[f].border = border
        ws[f].alignment = wrap_left_alignment

    file_name = 'temp_files/regresult{}.xlsx'.format(str(uuid4()))
    wb.save(file_name)

    task.file_path = file_name
    task.save()


class RegisterStatisticsExcelView(generics.RetrieveAPIView):
    """Создать запрос на Статистику регистрации Excel
    study_year(!), reg_period(!), acad_period, faculty, speciality, edu_prog, course, group
    """
    queryset = org_models.StudentDiscipline.objects.filter(is_active=True)

    def get(self, request, *args, **kwargs):
        profile = request.user.profile

        study_year = request.query_params.get('study_year')
        reg_period = request.query_params.get('reg_period')
        acad_period = request.query_params.get('acad_period')
        faculty = request.query_params.get('faculty')
        speciality = request.query_params.get('speciality')
        edu_prog = request.query_params.get('edu_prog')
        course = request.query_params.get('course')
        group = request.query_params.get('group')

        fields = {
            'study_year': study_year,
            'reg_period': reg_period,
            'acad_period': acad_period,
            'faculty': faculty,
            'speciality': speciality,
            'edu_prog': edu_prog,
            'course': course,
            'group': group,
        }
        token = uuid4()
        fields_json = json.dumps(fields)
        ExcelTask.objects.create(
            doc_type=2,
            profile=profile,
            token=token,
            fields=fields_json,
        )
        return Response(
            {
                'token': str(token)
            },
            status=status.HTTP_200_OK
        )


def make_register_statistics_excel(task):
    """Создает Статистику регистрации Excel
    study_year(!), reg_period(!), acad_period, faculty, speciality, edu_prog, course, group
    """
    queryset = org_models.StudentDiscipline.objects.filter(is_active=True)
    queryset = queryset.exclude(student__status_id=STUDENT_STATUSES['expelled'])

    profile = task.profile
    fields_json = task.fields
    fields = json.loads(fields_json)

    study_year = fields.get('study_year')
    reg_period = fields.get('reg_period')
    acad_period = fields.get('acad_period')
    faculty = fields.get('faculty')
    speciality = fields.get('speciality')
    edu_prog = fields.get('edu_prog')
    course = fields.get('course')
    group = fields.get('group')

    queryset = queryset.filter(
        status_id=student_discipline_status['not_chosen'],
        study_plan__advisor=profile,
    )

    wb = load_workbook('advisors/excel/register_stat.xlsx')
    ws = wb.active
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))
    font = Font(
        name='Times New Roman',
        size=11
    )
    font_small = Font(
        name='Times New Roman',
        size=8,
    )
    alignment = Alignment(
        vertical="center",
        horizontal="center"
    )

    wrap_left_alignment = Alignment(
        wrapText=True,
        vertical="center",
        horizontal="left",
    )

    if reg_period:
        reg_period_obj = common_models.RegistrationPeriod.objects.get(pk=reg_period)
        ws['B6'] = reg_period_obj.name
        ws['B6'].font = font_small
        ws['B6'].alignment = wrap_left_alignment

    if acad_period:
        queryset = queryset.filter(acad_period_id=acad_period)
        acad_period_obj = org_models.AcadPeriod.objects.get(pk=acad_period)
        ws['B7'] = acad_period_obj.repr_name
        ws['B7'].alignment = wrap_left_alignment
    else:
        ws['B7'] = 'Все'

        acad_period_pks = common_models.CourseAcadPeriodPermission.objects.filter(
            registration_period_id=reg_period,
            # course=course
        ).values('acad_period')
        queryset = queryset.filter(acad_period__in=acad_period_pks)

    ws['B7'].font = font_small

    if faculty:
        queryset = queryset.filter(study_plan__faculty_id=faculty)
        faculty_obj = org_models.Faculty.objects.get(pk=faculty)
        ws['B8'] = faculty_obj.name
        ws['B8'].alignment = wrap_left_alignment
    else:
        ws['B8'] = 'Все'
    ws['B8'].font = font_small

    if speciality:
        queryset = queryset.filter(study_plan__speciality_id=speciality)
        speciality_obj = org_models.Speciality.objects.get(pk=speciality)
        ws['B9'] = speciality_obj.name
        ws['B9'].alignment = wrap_left_alignment
    else:
        ws['B9'] = 'Все'
    ws['B9'].font = font_small

    if edu_prog:
        queryset = queryset.filter(study_plan__education_program_id=edu_prog)
        edu_prog_obj = org_models.EducationProgram.objects.get(pk=edu_prog)
        ws['B10'] = edu_prog_obj.name
        ws['B10'].alignment = wrap_left_alignment
    else:
        ws['B10'] = 'Все'
    ws['B10'].font = font_small

    if course:
        ws['B11'] = course
    else:
        ws['B11'] = 'Все'
    ws['B11'].font = font_small

    if group:
        queryset = queryset.filter(study_plan__group_id=group)
        group_obj = org_models.Group.objects.get(pk=group)
        ws['B12'] = group_obj.name
        ws['B12'].alignment = wrap_left_alignment
    else:
        ws['B12'] = 'Все'
    ws['B12'].font = font_small

    now = datetime.now()
    ws['B13'] = now.strftime("%d:%m:%Y, %H:%M:%S")
    ws['B13'].font = font_small

    if study_year:
        study_year_obj = org_models.StudyPeriod.objects.get(pk=study_year)
        ws['B5'] = '{} - {}'.format(study_year_obj.start,
                                    study_year_obj.end)
        ws['B5'].font = font_small
        ws['B5'].alignment = wrap_left_alignment

        queryset = queryset.filter(study_year_id=study_year)

    if course and study_year:
        study_plan_pks = org_models.StudyYearCourse.objects.filter(
            study_year_id=study_year,
            course=course
        ).values('study_plan')
        queryset = queryset.filter(study_plan__in=study_plan_pks)

    distincted_queryset = queryset.distinct('discipline', 'study_plan__group')

    student_discipline_list = []
    for student_discipline in distincted_queryset:
        group_student_count = org_models.StudyPlan.objects.filter(
            group=student_discipline.study_plan.group,
            is_active=True,
        ).distinct('student').count()

        not_chosen_student_count = queryset.filter(
            study_plan__group=student_discipline.study_plan.group,
            discipline=student_discipline.discipline
        ).distinct('student').count()

        d = {
            'faculty': student_discipline.study_plan.faculty.name,
            'cathedra': student_discipline.study_plan.cathedra.name,
            'speciality': student_discipline.study_plan.speciality.name,
            'group': student_discipline.study_plan.group.name,
            'student_count': group_student_count,
            'discipline': student_discipline.discipline.name,
            'not_chosen_student_count': not_chosen_student_count,
            'percent_of_non_chosen_student': (not_chosen_student_count / group_student_count) * 100,
        }
        student_discipline_list.append(d)

    for i, sd in enumerate(student_discipline_list):
        row_num = str(15 + i)
        a = 'A' + row_num
        ws[a] = sd['faculty']
        ws[a].font = font
        ws[a].border = border
        ws[a].alignment = wrap_left_alignment

        b = 'B' + row_num
        ws[b] = sd['cathedra']
        ws[b].font = font
        ws[b].border = border
        ws[b].alignment = wrap_left_alignment

        c = 'C' + row_num
        ws[c] = sd['speciality']
        ws[c].font = font
        ws[c].border = border
        ws[c].alignment = wrap_left_alignment

        d = 'D' + row_num
        ws[d] = sd['group']
        ws[d].font = font
        ws[d].border = border
        ws[d].alignment = wrap_left_alignment

        e = 'E' + row_num
        ws[e] = sd['student_count']
        ws[e].font = font
        ws[e].border = border
        ws[e].alignment = wrap_left_alignment

        f = 'F' + row_num
        ws[f] = sd['discipline']
        ws[f].font = font
        ws[f].border = border
        ws[f].alignment = wrap_left_alignment

        g = 'G' + row_num
        ws[g] = sd['not_chosen_student_count']
        ws[g].font = font
        ws[g].border = border
        ws[g].alignment = wrap_left_alignment

        h = 'H' + row_num
        ws[h] = sd['percent_of_non_chosen_student']
        ws[h].font = font
        ws[h].border = border
        ws[h].alignment = wrap_left_alignment

    file_name = 'temp_files/register_stat{}.xlsx'.format(str(uuid4()))
    wb.save(file_name)
    task.file_path = file_name
    task.save()

    # with open(file_name, 'rb') as f:
    #     response = HttpResponse(f, content_type='application/ms-excel')
    #     response['Content-Disposition'] = 'attachment; filename="reg_stat' + str(uuid4()) + '.xlsx"'
    #     return response


class NotRegisteredStudentListExcelView(generics.RetrieveAPIView):
    """Создать запрос на Список незарегистрированных Excel
    study_year(!), reg_period(!), acad_period, faculty, speciality, edu_prog, course, group
    """

    def get(self, request, *args, **kwargs):
        profile = self.request.user.profile

        study_year = self.request.query_params.get('study_year')
        reg_period = self.request.query_params.get('reg_period')
        acad_period = self.request.query_params.get('acad_period')
        faculty = self.request.query_params.get('faculty')
        speciality = self.request.query_params.get('speciality')
        edu_prog = self.request.query_params.get('edu_prog')
        course = self.request.query_params.get('course')
        group = self.request.query_params.get('group')

        fields = {
            'study_year': study_year,
            'reg_period': reg_period,
            'acad_period': acad_period,
            'faculty': faculty,
            'speciality': speciality,
            'edu_prog': edu_prog,
            'course': course,
            'group': group,
        }
        token = uuid4()
        fields_json = json.dumps(fields)
        ExcelTask.objects.create(
            doc_type=3,
            profile=profile,
            token=token,
            fields=fields_json,
        )
        return Response(
            {
                'token': str(token)
            },
            status=status.HTTP_200_OK
        )


def make_not_registered_student_excel(task):
    """Список незарегистрированных Excel
    study_year(!), reg_period(!), acad_period, faculty, speciality, edu_prog, course, group
    """
    queryset = org_models.StudentDiscipline.objects.filter(is_active=True)
    queryset = queryset.exclude(student__status_id=STUDENT_STATUSES['expelled'])

    profile = task.profile
    fields_json = task.fields
    fields = json.loads(fields_json)

    study_year = fields.get('study_year')
    reg_period = fields.get('reg_period')
    acad_period = fields.get('acad_period')
    faculty = fields.get('faculty')
    speciality = fields.get('speciality')
    edu_prog = fields.get('edu_prog')
    course = fields.get('course')
    group = fields.get('group')

    queryset = queryset.filter(
        status_id=student_discipline_status['not_chosen'],
        study_plan__advisor=profile,
    ).distinct('student')

    wb = load_workbook('advisors/excel/not_registered_students.xlsx')
    ws = wb.active
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))
    font = Font(
        name='Times New Roman',
        size=11
    )
    font_small = Font(
        name='Times New Roman',
        size=8,
    )

    wrap_left_alignment = Alignment(
        wrapText=True,
        vertical="center",
        horizontal="left",
    )

    if reg_period:
        reg_period_obj = common_models.RegistrationPeriod.objects.get(pk=reg_period)
        ws['B6'] = reg_period_obj.name
        ws['B6'].font = font_small
        ws['B6'].alignment = wrap_left_alignment

    if acad_period:
        queryset = queryset.filter(acad_period_id=acad_period)

        acad_period_obj = org_models.AcadPeriod.objects.get(pk=acad_period)
        ws['B7'] = acad_period_obj.repr_name
        ws['B7'].alignment = wrap_left_alignment
    else:
        ws['B7'] = 'Все'

        acad_period_pks = common_models.CourseAcadPeriodPermission.objects.filter(
            registration_period_id=reg_period,
            # course=course
        ).values('acad_period')
        queryset = queryset.filter(acad_period__in=acad_period_pks)

    ws['B7'].font = font_small

    if faculty:
        queryset = queryset.filter(study_plan__faculty_id=faculty)

        faculty_obj = org_models.Faculty.objects.get(pk=faculty)
        ws['B8'] = faculty_obj.name
        ws['B8'].alignment = wrap_left_alignment
    else:
        ws['B8'] = 'Все'
    ws['B8'].font = font_small

    if speciality:
        queryset = queryset.filter(study_plan__speciality_id=speciality)

        speciality_obj = org_models.Speciality.objects.get(pk=speciality)
        ws['B9'] = speciality_obj.name
        ws['B9'].alignment = wrap_left_alignment
    else:
        ws['B9'] = 'Все'
    ws['B9'].font = font_small

    if edu_prog:
        queryset = queryset.filter(study_plan__education_program_id=edu_prog)

        edu_prog_obj = org_models.EducationProgram.objects.get(pk=edu_prog)
        ws['B10'] = edu_prog_obj.name
        ws['B10'].alignment = wrap_left_alignment
    else:
        ws['B10'] = 'Все'
    ws['B10'].font = font_small

    if course:
        ws['B11'] = course
    else:
        ws['B11'] = 'Все'
    ws['B11'].font = font_small

    if group:
        queryset = queryset.filter(study_plan__group_id=group)

        group_obj = org_models.Group.objects.get(pk=group)
        ws['B12'] = group_obj.name
        ws['B12'].alignment = wrap_left_alignment
    else:
        ws['B12'] = 'Все'
    ws['B12'].font = font_small

    now = datetime.now()
    ws['B13'] = now.strftime("%d:%m:%Y, %H:%M:%S")
    ws['B13'].font = font_small

    if study_year:
        study_year_obj = org_models.StudyPeriod.objects.get(pk=study_year)
        ws['B5'] = '{} - {}'.format(study_year_obj.start,
                                    study_year_obj.end)
        ws['B5'].font = font_small
        ws['B5'].alignment = wrap_left_alignment

        queryset = queryset.filter(study_year_id=study_year)

    if course and study_year:
        study_plan_pks = org_models.StudyYearCourse.objects.filter(
            study_year_id=study_year,
            course=course
        ).values('study_plan')
        queryset = queryset.filter(study_plan__in=study_plan_pks)

    distincted_queryset = queryset.distinct(
        'study_plan__faculty',
        'study_plan__cathedra',
        'study_plan__speciality',
        'study_plan__group',
        'discipline',
    )
    student_discipline_list = []
    for item in distincted_queryset:
        sds = queryset.filter(
            study_plan__faculty=item.study_plan.faculty,
            study_plan__cathedra=item.study_plan.cathedra,
            study_plan__speciality=item.study_plan.speciality,
            study_plan__group=item.study_plan.group,
            discipline=item.discipline
        )
        student_name_list = [sd.study_plan.student.full_name.strip() for sd in sds]
        student_names = ', '.join(student_name_list)

        d = {
            'faculty': item.study_plan.faculty.name,
            'cathedra': item.study_plan.cathedra.name,
            'speciality': item.study_plan.speciality.name,
            'group': item.study_plan.group.name,
            'discipline': item.discipline.name,
            'student': student_names,
        }
        student_discipline_list.append(d)

    for i, sd in enumerate(student_discipline_list):
        row_num = str(15 + i)
        a = 'A' + row_num
        ws[a] = sd['faculty']
        ws[a].font = font
        ws[a].border = border
        ws[a].alignment = wrap_left_alignment

        b = 'B' + row_num
        ws[b] = sd['cathedra']
        ws[b].font = font
        ws[b].border = border
        ws[b].alignment = wrap_left_alignment

        c = 'C' + row_num
        ws[c] = sd['speciality']
        ws[c].font = font
        ws[c].border = border
        ws[c].alignment = wrap_left_alignment

        d = 'D' + row_num
        ws[d] = sd['group']
        ws[d].font = font
        ws[d].border = border
        ws[d].alignment = wrap_left_alignment

        e = 'E' + row_num
        ws[e] = sd['discipline']
        ws[e].font = font
        ws[e].border = border
        ws[e].alignment = wrap_left_alignment

        f = 'F' + row_num
        ws[f] = sd['student']
        ws[f].font = font
        ws[f].border = border
        ws[f].alignment = wrap_left_alignment

    file_name = 'temp_files/not_registered{}.xlsx'.format(str(uuid4()))
    wb.save(file_name)
    task.file_path = file_name
    task.save()


class GetFileView(generics.RetrieveAPIView):
    """Получить готовый файл"""
    permission_classes = (
        IsAuthenticated,
        permissions.ExcelAuthorPermission,
    )

    def get(self, request, *args, **kwargs):
        token = request.query_params.get('token')
        get = request.query_params.get('get')

        try:
            task = ExcelTask.objects.get(token=token,
                                         is_success=True,
                                         is_active=True)
            self.check_object_permissions(request,
                                          task)

            if get == 'status':
                resp = {
                    'done': True
                }
            elif get == 'file':
                with open(task.file_path, 'rb') as f:
                    response = HttpResponse(f, content_type='application/ms-excel')
                    response['Content-Disposition'] = 'attachment; filename="regresult' + str(uuid4()) + '.xls"'
                    return response
            else:
                return Response(
                    {
                        'message': 'invalid_param',
                    },
                    status=status.HTTP_200_OK
                )
        except ExcelTask.DoesNotExist:
            if get == 'status':
                resp = {
                    'done': False
                }
            elif get == 'file':
                return Response(
                    {
                        'message': 'not_found',
                    },
                    status=status.HTTP_200_OK,
                )
            else:
                return Response(
                    {
                        'message': 'invalid_param',
                    },
                    status=status.HTTP_200_OK
                )
        return Response(
            resp,
            status=status.HTTP_200_OK
        )
